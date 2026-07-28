import asyncio
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from PIL import Image
from starlette.datastructures import UploadFile

from backend.api.prelogs import (
    export_prelog,
    filter_playlist_files,
    inspect_playlist_file,
    playlist_filter_options,
)
from backend.api.postlogs import export_postlog, filter_postlog_events
from backend.services.traffic.prelog_export import generate_prelog_workbook
from openpyxl import Workbook, load_workbook
from backend.services.traffic.playlist import (
    filter_playlist_events,
    inspect_playlist,
    parse_playlist_file,
    parse_playlist_events,
)


SAMPLE_PLAYLIST = Path("tests/sample_playlist.csv")


def test_playlist_inspection_uses_embedded_metadata():
    result = inspect_playlist(SAMPLE_PLAYLIST.read_bytes())

    assert result["metadata"]["date"] == "2026-07-25"
    assert result["metadata"]["start_time"] == "06:00:00"
    assert result["metadata"]["channel_name"] == "Comercio TV"
    assert result["header_row"] == 2


def test_playlist_inspection_detects_columns_and_assets():
    result = inspect_playlist(SAMPLE_PLAYLIST.read_bytes())

    assert result["detected_columns"] == {
        "asset_id": "ASSET ID",
        "time": "HOUR",
        "duration": "DURATION",
    }
    assert result["rows"] == 5
    assert result["asset_occurrences"] == 5
    assert result["unique_assets"] == 4
    assert result["prefixes"][0] == {
        "prefix": "promo_",
        "occurrences": 2,
    }


def test_playlist_endpoint_ignores_filename():
    upload = UploadFile(
        filename="meaningless-name.csv",
        file=BytesIO(SAMPLE_PLAYLIST.read_bytes()),
    )
    result = asyncio.run(inspect_playlist_file(upload))

    assert result["filename"] == "meaningless-name.csv"
    assert result["metadata"]["channel_name"] == "Comercio TV"
    assert result["metadata"]["date"] == "2026-07-25"


def test_playlist_requires_detectable_or_mapped_headers():
    try:
        inspect_playlist(b"one,two,three\n1,2,3\n")
    except ValueError as exc:
        assert "Manual column mapping" in str(exc)
    else:
        raise AssertionError("Expected an unmapped playlist to be rejected.")


def test_playlist_rolls_events_into_the_next_day():
    content = SAMPLE_PLAYLIST.read_text().replace(
        "06:00:00,0:00:30,promo_open",
        "23:59:30,0:00:30,promo_open",
    ).replace(
        "06:00:30,0:00:45,client_spot_a",
        "00:00:00,0:00:45,client_spot_a",
    )
    _, events = parse_playlist_events(content.encode())

    assert events[2].air_datetime.isoformat() == "2026-07-26T00:00:00"


def test_small_time_regression_does_not_advance_the_date():
    content = SAMPLE_PLAYLIST.read_text().replace(
        "06:01:15,0:00:45,client_spot_a",
        "05:59:00,0:00:45,client_spot_a",
    )
    _, events = parse_playlist_events(content.encode())

    assert events[3].air_datetime.date().isoformat() == "2026-07-25"


def test_twelve_hour_playlist_cycles_infer_am_pm_and_next_day():
    content = b"""2026-07-26,06:00:00,,Tarima TV
HOUR,DURATION,ASSET ID
11:59:00,0:00:30,morning
12:00:00,0:00:30,noon
01:00:00,0:00:30,afternoon
11:59:00,0:00:30,night
12:00:00,0:00:30,midnight
01:00:00,0:00:30,overnight
"""
    _, events = parse_playlist_events(content)

    assert events[2].air_datetime.isoformat() == "2026-07-26T13:00:00"
    assert events[4].air_datetime.isoformat() == "2026-07-27T00:00:00"
    assert events[5].air_datetime.isoformat() == "2026-07-27T01:00:00"


def test_filter_builder_supports_prefix_exact_and_contains():
    _, events = parse_playlist_events(SAMPLE_PLAYLIST.read_bytes())

    prefix_matches = filter_playlist_events(
        events,
        filter_mode="prefix",
        filter_value="promo_",
    )
    exact_matches = filter_playlist_events(
        events,
        filter_mode="exact",
        filter_value="client_spot_a",
    )
    contains_matches = filter_playlist_events(
        events,
        filter_mode="contains",
        filter_value="spot",
    )

    assert len(prefix_matches) == 2
    assert len(exact_matches) == 2
    assert len(contains_matches) == 2


def test_filter_builder_supports_date_and_overnight_time_ranges():
    _, events = parse_playlist_events(SAMPLE_PLAYLIST.read_bytes())
    matches = filter_playlist_events(
        events,
        start_date="2026-07-25",
        end_date="2026-07-25",
        start_time="06:00:30",
        end_time="06:01:15",
    )

    assert [event.asset_id for event in matches] == [
        "client_spot_a",
        "client_spot_a",
    ]


def test_filter_endpoint_combines_multiple_files():
    uploads = [
        UploadFile(
            filename=f"playlist-{position}.csv",
            file=BytesIO(SAMPLE_PLAYLIST.read_bytes()),
        )
        for position in range(2)
    ]
    result = asyncio.run(
        filter_playlist_files(
            uploads,
            "prefix",
            "promo_",
            None,
            None,
            None,
            None,
            "06:00:00",
            None,
        )
    )

    assert result["files_processed"] == 2
    assert result["events_received"] == 10
    assert result["matching_events"] == 4
    assert result["unique_assets"] == 2


def test_filter_options_combine_assets_and_dates():
    uploads = [
        UploadFile(
            filename="anything.csv",
            file=BytesIO(SAMPLE_PLAYLIST.read_bytes()),
        )
    ]
    result = asyncio.run(playlist_filter_options(uploads, None))

    assert result["channels"] == ["Comercio TV"]
    assert result["start_date"] == "2026-07-25"
    assert result["end_date"] == "2026-07-25"
    assert result["assets"][0]["asset_id"] == "Morning Programming"
    assert any(
        item["prefix"] == "promo_"
        for item in result["prefixes"]
    )


def test_broadcast_date_includes_overnight_events_until_next_six_am():
    content = b"""2026-07-26,06:00:00,,Tarima TV
HOUR,DURATION,ASSET ID
11:59:00,0:00:30,day
12:00:00,0:00:30,noon
01:00:00,0:00:30,afternoon
11:59:00,0:00:30,night
12:00:00,0:00:30,midnight
05:59:59,0:00:30,last_event
06:00:00,0:00:30,next_broadcast_day
"""
    _, events = parse_playlist_events(content)
    matches = filter_playlist_events(
        events,
        start_date="2026-07-26",
        end_date="2026-07-26",
        broadcast_day_start="06:00:00",
    )

    assert matches[-1].asset_id == "last_event"
    assert "next_broadcast_day" not in {
        event.asset_id
        for event in matches
    }


def test_auto_timezone_requires_playlist_metadata():
    try:
        parse_playlist_events(
            SAMPLE_PLAYLIST.read_bytes(),
            source_timezone="auto",
        )
    except ValueError as exc:
        assert "does not declare a time zone" in str(exc)
    else:
        raise AssertionError("Expected manual time zone selection.")


def test_selected_timezone_is_attached_to_events():
    _, events = parse_playlist_events(
        SAMPLE_PLAYLIST.read_bytes(),
        source_timezone="America/New_York",
    )

    assert events[0].air_datetime.utcoffset() is not None
    assert events[0].air_datetime.isoformat().endswith("-04:00")


def test_proteus_utf16_fixed_width_asrun_is_detected():
    def proteus_row(
        asset_id: str,
        air_date: str,
        timecode: str,
        duration: str,
    ) -> str:
        return (
            f"{asset_id:<38}"
            f"{'PROMOCIONES/TEST':<28}"
            f"{'Aired':<13}"
            f"{air_date:<14}"
            f"{timecode:<15}"
            f"{duration:<11}"
            f"{'':<20}"
            f"{'PROMOCIONES/TEST':<19}"
        )

    content = "\r\n".join([
        proteus_row(
            "HPR9660US",
            "10/28/24",
            "06:11:23;27",
            "00:00:33;15",
        ),
        proteus_row(
            "HPR9893US",
            "10/28/24",
            "06:11:57;12",
            "00:00:33;22",
        ),
    ]).encode("utf-16-le")

    structure, events = parse_playlist_file(
        "unrelated-name.txt",
        content,
        "America/New_York",
    )

    assert structure["format"] == "Fixed-width As-Run"
    assert structure["encoding"] == "Auto-detected"
    assert structure["asset_occurrences"] == 2
    assert events[0].asset_id == "HPR9660US"
    assert events[0].air_datetime.isoformat() == (
        "2024-10-28T06:11:23-04:00"
    )
    assert events[0].duration == "00:00:33;15"


def test_fixed_width_asrun_detection_does_not_require_vendor_offsets():
    content = "\n".join([
        "Report generated by an unrelated playout system",
        (
            "SPOT-42      Client spot title     Played       "
            "2024-10-29     07:15:00.12     00:00:30.00"
        ),
        (
            "PROMO-A-7        Station promo with a longer description"
            "        TX OK     2024-10-29   08:45:10;00   00:01:00;00"
        ),
    ]).encode("utf-16-le")

    structure, events = parse_playlist_file(
        "vendor-neutral.txt",
        content,
        "America/New_York",
    )

    assert structure["format"] == "Fixed-width As-Run"
    assert structure["asset_occurrences"] == 2
    assert [event.asset_id for event in events] == [
        "SPOT-42",
        "PROMO-A-7",
    ]
    assert events[1].air_datetime.isoformat() == (
        "2024-10-29T08:45:10-04:00"
    )


def test_xml_playlist_uses_common_item_fields_and_rolls_after_midnight():
    content = b"""<?xml version="1.0" encoding="UTF-8"?>
<traffics>
  <traffic channelid="TEST">
    <item mediaid="PROMO-1">
      <startat>23:59:45;00</startat>
      <duration>00:00:15;00</duration>
    </item>
    <item mediaid="SPOT-2">
      <startat>00:00:00;00</startat>
      <duration>00:00:30;00</duration>
    </item>
  </traffic>
</traffics>"""

    structure, events = parse_playlist_file(
        "name-without-a-date.xml",
        content,
        "America/New_York",
        operational_date="2026-07-27",
    )

    assert structure["format"] == "XML playlist"
    assert structure["metadata"]["channel_name"] == "TEST"
    assert [event.asset_id for event in events] == ["PROMO-1", "SPOT-2"]
    assert events[0].air_datetime.isoformat() == (
        "2026-07-27T23:59:45-04:00"
    )
    assert events[1].air_datetime.isoformat() == (
        "2026-07-28T00:00:00-04:00"
    )


def test_tab_delimited_spanish_playlist_headers_are_detected():
    content = (
        "N.Ord.	Hora	Tipo_Even.	ID_Cinta	Duracion\n"
        "1	23:59:45:00	PROM	HPR100#16:9	00:00:15:00\n"
        "2	00:00:00:00	PROM	HPR200#	00:00:30:00\n"
    ).encode()

    structure, events = parse_playlist_file(
        "schedule-without-date.txt",
        content,
        "America/New_York",
        operational_date="2026-07-27",
    )

    assert structure["detected_columns"] == {
        "asset_id": "ID_Cinta",
        "time": "Hora",
        "duration": "Duracion",
    }
    assert [event.asset_id for event in events] == ["HPR100", "HPR200"]
    assert events[0].air_datetime.isoformat() == (
        "2026-07-27T23:59:45-04:00"
    )
    assert events[1].air_datetime.isoformat() == (
        "2026-07-28T00:00:00-04:00"
    )


def test_prelog_workbook_uses_requested_language_and_columns():
    _, events = parse_playlist_events(SAMPLE_PLAYLIST.read_bytes())
    content = generate_prelog_workbook(
        events[:2],
        channel_name="Comercio TV",
        language="es",
        product="Campaña institucional",
    )
    workbook = load_workbook(BytesIO(content))
    worksheet = workbook["Pre Log"]

    assert worksheet["A5"].value == "Nombre del canal"
    assert worksheet["B5"].value == "Producto"
    assert worksheet["C5"].value == "Identificador del elemento"
    assert worksheet["D5"].value == "Fecha"
    assert worksheet["E5"].value == "Hora"
    assert worksheet["F5"].value == "Duración"
    assert worksheet["A6"].value == "Comercio TV"
    assert worksheet["B6"].value == "Campaña institucional"
    assert worksheet["D6"].number_format == "mm/dd/yyyy"
    assert worksheet["E6"].number_format == "hh:mm:ss"
    assert "Pago" not in {
        cell.value
        for cell in worksheet[5]
    }


def test_prelog_export_endpoint_downloads_xlsx():
    upload = UploadFile(
        filename="ignored.csv",
        file=BytesIO(SAMPLE_PLAYLIST.read_bytes()),
    )
    response = asyncio.run(export_prelog(
        [upload],
        "prefix",
        "promo_",
        "2026-07-25",
        "2026-07-25",
        "06:00:00",
        "America/New_York",
        "Comercio TV",
        "en",
        None,
        None,
        None,
        "xlsx",
    ))

    assert response.media_type.endswith("spreadsheetml.sheet")
    workbook = load_workbook(BytesIO(response.body))
    assert workbook["Pre Log"]["A6"].value == "Comercio TV"


def test_prelog_workbook_embeds_jpg_logo():
    _, events = parse_playlist_events(SAMPLE_PLAYLIST.read_bytes())
    logo = BytesIO()
    Image.new("RGB", (160, 60), "#808080").save(logo, format="JPEG")
    content = generate_prelog_workbook(
        events[:2],
        channel_name="Tarima TV",
        logo_content=logo.getvalue(),
    )

    with ZipFile(BytesIO(content)) as archive:
        media_files = [
            name
            for name in archive.namelist()
            if name.startswith("xl/media/")
        ]

    assert len(media_files) == 1


def test_postlog_filters_actual_airings():
    upload = UploadFile(
        filename="as-run.csv",
        file=BytesIO(SAMPLE_PLAYLIST.read_bytes()),
    )
    result = asyncio.run(filter_postlog_events(
        [upload],
        "exact",
        "client_spot_a",
        "2026-07-25",
        "2026-07-25",
        "06:00:00",
        "America/New_York",
    ))

    assert result["matching_events"] == 2
    assert result["unique_assets"] == 1


def test_postlog_export_is_a_broadcast_certification():
    upload = UploadFile(
        filename="as-run.csv",
        file=BytesIO(SAMPLE_PLAYLIST.read_bytes()),
    )
    response = asyncio.run(export_postlog(
        [upload],
        "exact",
        "client_spot_a",
        "2026-07-25",
        "2026-07-25",
        "06:00:00",
        "America/New_York",
        "Comercio TV",
        "en",
        "Campaign A",
        None,
        None,
        "xlsx",
    ))
    workbook = load_workbook(BytesIO(response.body))
    worksheet = workbook["Pre Log"]

    assert worksheet["C1"].value == "Post Log — Broadcast Certification"
    assert worksheet["B5"].value == "Product"
    assert "Total Airings: 2" in worksheet["A9"].value
    assert "postlog-comercio-tv" in response.headers["content-disposition"]


def test_postlog_export_separates_each_asset_into_its_own_workbook():
    upload = UploadFile(
        filename="as-run.csv",
        file=BytesIO(SAMPLE_PLAYLIST.read_bytes()),
    )
    response = asyncio.run(export_postlog(
        [upload],
        "prefix",
        "promo_",
        "2026-07-25",
        "2026-07-25",
        "06:00:00",
        "America/New_York",
        "Comercio TV",
        "en",
        None,
        None,
        None,
        "xlsx",
    ))

    assert response.media_type == "application/zip"
    with ZipFile(BytesIO(response.body)) as archive:
        names = archive.namelist()
        workbooks = [
            load_workbook(BytesIO(archive.read(name)))
            for name in names
        ]

    assert len(names) == 2
    assert any("promo-open" in name for name in names)
    assert any("promo-close" in name for name in names)
    assert all(
        len({
            worksheet.cell(row, 2).value
            for row in range(6, worksheet.max_row)
            if worksheet.cell(row, 2).value
        }) == 1
        for workbook in workbooks
        for worksheet in [workbook["Pre Log"]]
    )


def test_prelog_and_postlog_support_pdf_downloads():
    prelog_upload = UploadFile(
        filename="playlist.csv",
        file=BytesIO(SAMPLE_PLAYLIST.read_bytes()),
    )
    prelog_response = asyncio.run(export_prelog(
        [prelog_upload],
        "exact",
        "client_spot_a",
        "2026-07-25",
        "2026-07-25",
        "06:00:00",
        "America/New_York",
        "Comercio TV",
        "en",
        None,
        None,
        None,
        "pdf",
    ))

    postlog_upload = UploadFile(
        filename="as-run.csv",
        file=BytesIO(SAMPLE_PLAYLIST.read_bytes()),
    )
    postlog_response = asyncio.run(export_postlog(
        [postlog_upload],
        "exact",
        "client_spot_a",
        "2026-07-25",
        "2026-07-25",
        "06:00:00",
        "America/New_York",
        "Comercio TV",
        "en",
        None,
        None,
        None,
        "pdf",
    ))

    assert prelog_response.media_type == "application/pdf"
    assert postlog_response.media_type == "application/pdf"
    assert prelog_response.body.startswith(b"%PDF")
    assert postlog_response.body.startswith(b"%PDF")


def test_amagi_excel_as_run_is_normalized():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Asrun Report"
    worksheet.append([
        "Playlist",
        "Item ID",
        "Asset Name",
        "Headend Start Time",
        "Duration Played",
    ])
    worksheet.append([
        4057,
        2,
        "ca_indotel",
        "2026-07-19 13:55:55",
        "00:00:45:01",
    ])
    content = BytesIO()
    workbook.save(content)

    structure, events = parse_playlist_file(
        "amagi-asrun.xlsx",
        content.getvalue(),
        "America/New_York",
    )

    assert structure["detected_columns"]["asset_id"] == "Asset Name"
    assert structure["detected_columns"]["time"] == "Headend Start Time"
    assert events[0].asset_id == "ca_indotel"
    assert events[0].air_datetime.isoformat() == (
        "2026-07-19T13:55:55-04:00"
    )
    assert events[0].duration == "00:00:45"


def test_json_txt_and_xml_as_runs_share_the_internal_model():
    records = [
        {
            "asset_id": "ca_indotel",
            "start_time": "2026-07-20 13:28:12",
            "duration": "00:00:45:01",
        }
    ]
    json_content = (
        b'{"events":[{"asset_id":"ca_indotel",'
        b'"start_time":"2026-07-20 13:28:12",'
        b'"duration":"00:00:45:01"}]}'
    )
    txt_content = (
        b"asset_id|start_time|duration\n"
        b"ca_indotel|2026-07-20 13:28:12|00:00:45:01\n"
    )
    xml_content = b"""<events><event>
<asset_id>ca_indotel</asset_id>
<start_time>2026-07-20 13:28:12</start_time>
<duration>00:00:45:01</duration>
</event></events>"""

    results = [
        parse_playlist_file(
            filename,
            content,
            "America/New_York",
        )[1][0]
        for filename, content in (
            ("asrun.json", json_content),
            ("asrun.txt", txt_content),
            ("asrun.xml", xml_content),
        )
    ]

    assert all(event.asset_id == records[0]["asset_id"] for event in results)
    assert all(event.duration == "00:00:45" for event in results)
    assert all(
        event.air_datetime.isoformat() == "2026-07-20T13:28:12-04:00"
        for event in results
    )
