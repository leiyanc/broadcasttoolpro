from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from backend.api.xmltv import (
    EXCEL_TEMPLATE,
    download_csv_template,
    download_excel_template,
)
from backend.services.xmltv.parser import EXPECTED_COLUMNS


def test_excel_template_has_expected_structure():
    workbook = load_workbook(
        BytesIO(EXCEL_TEMPLATE.read_bytes()),
        read_only=True,
        data_only=True,
    )

    assert workbook.sheetnames == [
        "Programming",
        "Instructions",
        "Field Reference",
        "Example",
    ]
    worksheet = workbook["Programming"]
    headers = [
        cell.value for cell in worksheet[4][:len(EXPECTED_COLUMNS)]
    ]
    assert headers == EXPECTED_COLUMNS

    instructions = workbook["Instructions"]
    instruction_text = " ".join(
        str(cell.value or "")
        for row in instructions.iter_rows()
        for cell in row
    )
    assert "VCHIP" in instruction_text
    assert "coincidir exactamente" in instruction_text
    assert "Enter TV-PG, not PG" in instruction_text
    assert "Escribe TV-PG, no PG" in instruction_text
    assert "copy, paste, or drag" in instruction_text

    reference = workbook["Field Reference"]
    reference_text = " ".join(
        str(cell.value or "")
        for row in reference.iter_rows()
        for cell in row
    )
    assert "free text" in reference_text.lower()
    assert "texto libre" in reference_text.lower()


def test_excel_template_endpoint_returns_download():
    response = download_excel_template()

    assert Path(response.path) == EXCEL_TEMPLATE
    assert response.filename == "Broadcast_Tool_Pro_XMLTV_Template.xlsx"
    assert response.headers["cache-control"] == "no-store, max-age=0"


def test_excel_template_keeps_parental_rating_as_free_text():
    workbook = load_workbook(BytesIO(EXCEL_TEMPLATE.read_bytes()))
    worksheet = workbook["Programming"]
    validations = {
        str(validation.sqref): validation.formula1
        for validation in worksheet.data_validations.dataValidation
    }

    assert "F5:F504" not in validations
    assert worksheet.max_column == len(EXPECTED_COLUMNS)
    assert not workbook.defined_names


def test_csv_template_endpoint_returns_expected_headers():
    response = download_csv_template()
    content = response.body.decode()

    assert content.splitlines()[0].split(",") == EXPECTED_COLUMNS
    assert "Broadcast_Tool_Pro_XMLTV_Template.csv" in response.headers[
        "content-disposition"
    ]
    assert response.headers["cache-control"] == "no-store, max-age=0"
