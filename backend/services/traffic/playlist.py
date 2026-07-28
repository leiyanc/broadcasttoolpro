import csv
import json
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime, time, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from re import match
from typing import Any
from xml.etree import ElementTree
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import load_workbook

MAX_PLAYLIST_SIZE = 20 * 1024 * 1024
HEADER_ALIASES = {
    "asset_id": {
        "asset id",
        "asset",
        "program name",
        "programme name",
        "event id",
        "clip id",
        "media id",
        "asset name",
    },
    "time": {
        "hour",
        "time",
        "start time",
        "air time",
        "headend start time",
    },
    "duration": {
        "duration",
        "chrono",
        "length",
        "duration played",
    },
}


@dataclass(frozen=True)
class PlaylistEvent:
    channel_name: str | None
    air_datetime: datetime
    duration: str | None
    asset_id: str
    source_row: int

    def to_dict(self) -> dict[str, Any]:
        result = asdict(self)
        result["air_datetime"] = self.air_datetime.isoformat()
        return result


def _normalized(value: str) -> str:
    return " ".join(value.strip().lower().replace("_", " ").split())


def _decode_csv(content: bytes) -> str:
    if not content:
        raise ValueError("The playlist file is empty.")

    if len(content) > MAX_PLAYLIST_SIZE:
        raise ValueError("The playlist file exceeds the 20 MB import limit.")

    try:
        return content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("The playlist CSV must use UTF-8 encoding.") from exc


def _decode_text(content: bytes) -> str:
    if not content:
        raise ValueError("The As-Run file is empty.")
    if len(content) > MAX_PLAYLIST_SIZE:
        raise ValueError("The As-Run file exceeds the 20 MB import limit.")
    encodings = (
        ("utf-16", content.startswith((b"\xff\xfe", b"\xfe\xff"))),
        ("utf-16-le", content.count(b"\x00") > len(content) // 4),
        ("utf-8-sig", True),
    )
    for encoding, applicable in encodings:
        if not applicable:
            continue
        try:
            return content.decode(encoding).lstrip("\ufeff")
        except UnicodeDecodeError:
            continue
    raise ValueError(
        "The TXT As-Run encoding could not be detected."
    )


def _header_score(row: list[str]) -> int:
    normalized = {_normalized(value) for value in row if value.strip()}
    return sum(
        any(alias in normalized for alias in aliases)
        for aliases in HEADER_ALIASES.values()
    )


def _detect_header(rows: list[list[str]]) -> int:
    candidates = [
        (_header_score(row), position)
        for position, row in enumerate(rows[:25])
    ]
    score, position = max(candidates, default=(0, 0))

    if score < 2:
        raise ValueError(
            "The playlist header could not be detected. "
            "Manual column mapping is required."
        )

    return position


def _detect_columns(headers: list[str]) -> dict[str, str | None]:
    detected: dict[str, str | None] = {
        "asset_id": None,
        "time": None,
        "duration": None,
    }

    for header in headers:
        normalized = _normalized(header)
        for field, aliases in HEADER_ALIASES.items():
            if detected[field] is None and normalized in aliases:
                detected[field] = header

    return detected


def _parse_date(value: str) -> str | None:
    text = value.strip()
    formats = (
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%d-%m-%Y",
    )

    for value_format in formats:
        try:
            return datetime.strptime(text, value_format).date().isoformat()
        except ValueError:
            continue

    return None


def _parse_time(value: str) -> str | None:
    text = value.strip()
    formats = (
        "%H:%M:%S",
        "%H:%M",
        "%I:%M:%S %p",
        "%I:%M %p",
    )

    for value_format in formats:
        try:
            return datetime.strptime(text, value_format).time().isoformat()
        except ValueError:
            continue

    return None


def _metadata(rows: list[list[str]]) -> dict[str, Any]:
    embedded_date = None
    embedded_time = None
    channel_name = None
    source_timezone = None

    for row in rows:
        for value in row:
            if not value.strip():
                continue

            if embedded_date is None:
                embedded_date = _parse_date(value)
                if embedded_date is not None:
                    continue

            if embedded_time is None:
                embedded_time = _parse_time(value)
                if embedded_time is not None:
                    continue

            if source_timezone is None:
                timezone_candidate = value.strip()
                if timezone_candidate.upper() in {"UTC", "GMT"}:
                    source_timezone = "UTC"
                    continue
                elif "/" in timezone_candidate:
                    try:
                        ZoneInfo(timezone_candidate)
                        source_timezone = timezone_candidate
                        continue
                    except ZoneInfoNotFoundError:
                        pass

            if channel_name is None:
                channel_name = value.strip()

    return {
        "date": embedded_date,
        "start_time": embedded_time,
        "channel_name": channel_name,
        "source_timezone": source_timezone,
    }


def _prefix(asset_id: str) -> str:
    result = match(r"^[A-Za-z]+_", asset_id)
    return result.group(0).lower() if result else "(no prefix)"


def _seconds(value: time) -> int:
    return (value.hour * 3600) + (value.minute * 60) + value.second


def _playlist_structure(content: bytes) -> dict[str, Any]:
    text = _decode_csv(content)
    rows = [
        [str(value).strip() for value in row]
        for row in csv.reader(StringIO(text))
    ]

    if not any(any(value for value in row) for row in rows):
        raise ValueError("The playlist does not contain any data.")

    header_index = _detect_header(rows)
    headers = rows[header_index]
    detected_columns = _detect_columns(headers)
    metadata = _metadata(rows[:header_index])
    data_rows = [
        row
        for row in rows[header_index + 1:]
        if any(value for value in row)
    ]

    asset_column = detected_columns["asset_id"]
    asset_index = headers.index(asset_column) if asset_column else None
    assets = [
        row[asset_index].strip()
        for row in data_rows
        if asset_index is not None
        and asset_index < len(row)
        and row[asset_index].strip()
    ]
    asset_counts = Counter(assets)
    prefix_counts = Counter(_prefix(asset) for asset in assets)

    sample_rows = [
        {
            header: row[position] if position < len(row) else ""
            for position, header in enumerate(headers)
            if header
        }
        for row in data_rows[:10]
    ]

    return {
        "header_row": header_index + 1,
        "headers": [header for header in headers if header],
        "detected_columns": detected_columns,
        "metadata": metadata,
        "rows": len(data_rows),
        "asset_occurrences": len(assets),
        "unique_assets": len(asset_counts),
        "assets": [
            {
                "asset_id": asset,
                "occurrences": count,
            }
            for asset, count in sorted(asset_counts.items())
        ],
        "prefixes": [
            {
                "prefix": prefix,
                "occurrences": count,
            }
            for prefix, count in prefix_counts.most_common()
        ],
        "sample_rows": sample_rows,
        "_raw_rows": data_rows,
        "_raw_headers": headers,
    }


def inspect_playlist(content: bytes) -> dict[str, Any]:
    result = _playlist_structure(content)
    result.pop("_raw_rows")
    result.pop("_raw_headers")
    return result


def parse_playlist_events(
    content: bytes,
    source_timezone: str | None = None,
) -> tuple[dict[str, Any], list[PlaylistEvent]]:
    structure = _playlist_structure(content)
    metadata = structure["metadata"]
    detected = structure["detected_columns"]
    headers = structure["_raw_headers"]

    if not metadata["date"]:
        raise ValueError(
            "The playlist does not contain an embedded date. "
            "A manual date is required."
        )

    if not detected["time"] or not detected["asset_id"]:
        raise ValueError(
            "Time and Asset ID columns must be mapped before filtering."
        )

    time_index = headers.index(detected["time"])
    asset_index = headers.index(detected["asset_id"])
    duration_index = (
        headers.index(detected["duration"])
        if detected["duration"]
        else None
    )
    if source_timezone == "auto":
        timezone_name = metadata["source_timezone"]
        if not timezone_name:
            raise ValueError(
                "The playlist does not declare a time zone. "
                "Select the source time zone manually."
            )
    else:
        timezone_name = source_timezone or metadata["source_timezone"]
    timezone_info = None

    if timezone_name:
        try:
            timezone_info = ZoneInfo(timezone_name)
        except ZoneInfoNotFoundError as exc:
            raise ValueError(
                f'Unknown source time zone "{timezone_name}".'
            ) from exc

    base_date = datetime.strptime(
        metadata["date"],
        "%Y-%m-%d",
    ).replace(tzinfo=timezone_info)
    previous_raw_seconds: int | None = None
    previous_absolute_seconds: int | None = None
    cycle_offset = 0
    events: list[PlaylistEvent] = []

    for offset, row in enumerate(structure["_raw_rows"], start=1):
        if time_index >= len(row) or asset_index >= len(row):
            continue

        parsed_time = _parse_time(row[time_index])
        asset_id = row[asset_index].strip()

        if not parsed_time or not asset_id:
            continue

        event_time = time.fromisoformat(parsed_time)
        raw_seconds = _seconds(event_time)
        absolute_seconds = raw_seconds + cycle_offset

        if (
            previous_raw_seconds is not None
            and previous_absolute_seconds is not None
            and previous_raw_seconds - raw_seconds > 6 * 3600
        ):
            while absolute_seconds < previous_absolute_seconds:
                cycle_offset += 12 * 3600
                absolute_seconds = raw_seconds + cycle_offset

        previous_raw_seconds = raw_seconds
        previous_absolute_seconds = absolute_seconds
        duration = (
            row[duration_index].strip() or None
            if duration_index is not None and duration_index < len(row)
            else None
        )
        events.append(PlaylistEvent(
            channel_name=metadata["channel_name"],
            air_datetime=base_date + timedelta(seconds=absolute_seconds),
            duration=duration,
            asset_id=asset_id,
            source_row=structure["header_row"] + offset,
        ))

    public_structure = {
        key: value
        for key, value in structure.items()
        if key not in {"_raw_rows", "_raw_headers"}
    }
    return public_structure, events


def _amagi_duration(value: Any) -> str | None:
    if value is None:
        return None

    text = str(value).strip()
    parts = text.split(":")
    if len(parts) >= 3:
        return ":".join(parts[:3])
    return text or None


def _excel_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(microsecond=0)

    text = str(value or "").strip()
    for value_format in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S:%f",
        "%m/%d/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, value_format).replace(
                microsecond=0
            )
        except ValueError:
            continue
    return None


def _timezone(source_timezone: str | None) -> ZoneInfo | None:
    if source_timezone == "auto":
        raise ValueError(
            "The As-Run does not declare a time zone. "
            "Select the source time zone manually."
        )
    if not source_timezone:
        return None
    try:
        return ZoneInfo(source_timezone)
    except ZoneInfoNotFoundError as exc:
        raise ValueError(
            f'Unknown source time zone "{source_timezone}".'
        ) from exc


def _structured_events(
    headers: list[str],
    records: list[dict[str, Any]],
    source_timezone: str | None,
) -> tuple[dict[str, Any], list[PlaylistEvent]]:
    detected = _detect_columns(headers)
    if not detected["asset_id"] or not detected["time"]:
        raise ValueError(
            "Asset ID and Start Time fields must be mapped."
        )

    timezone_info = _timezone(source_timezone)
    channel_header = next(
        (
            header
            for header in headers
            if _normalized(header) in {"channel", "channel name"}
        ),
        None,
    )
    events = []
    for row_number, record in enumerate(records, start=2):
        asset_id = str(record.get(detected["asset_id"]) or "").strip()
        air_datetime = _excel_datetime(record.get(detected["time"]))
        if not asset_id or air_datetime is None:
            continue
        events.append(PlaylistEvent(
            channel_name=(
                str(record.get(channel_header) or "").strip() or None
                if channel_header
                else None
            ),
            air_datetime=air_datetime.replace(tzinfo=timezone_info),
            duration=_amagi_duration(
                record.get(detected["duration"])
                if detected["duration"]
                else None
            ),
            asset_id=asset_id,
            source_row=row_number,
        ))

    structure = {
        "header_row": 1,
        "headers": headers,
        "detected_columns": detected,
        "metadata": {
            "date": (
                min(event.air_datetime for event in events).date().isoformat()
                if events
                else None
            ),
            "start_time": None,
            "channel_name": next(
                (
                    event.channel_name
                    for event in events
                    if event.channel_name
                ),
                None,
            ),
            "source_timezone": source_timezone,
        },
        "rows": len(records),
        "asset_occurrences": len(events),
        "unique_assets": len({event.asset_id for event in events}),
    }
    return structure, events


def parse_json_playlist_events(
    content: bytes,
    source_timezone: str | None = None,
) -> tuple[dict[str, Any], list[PlaylistEvent]]:
    try:
        payload = json.loads(_decode_csv(content))
    except json.JSONDecodeError as exc:
        raise ValueError("The JSON As-Run file is invalid.") from exc

    if isinstance(payload, dict):
        records = next(
            (
                value
                for value in payload.values()
                if isinstance(value, list)
            ),
            None,
        )
    else:
        records = payload
    if not isinstance(records, list) or not all(
        isinstance(record, dict)
        for record in records
    ):
        raise ValueError(
            "The JSON As-Run must contain a list of event objects."
        )
    headers = list(dict.fromkeys(
        key
        for record in records
        for key in record
    ))
    return _structured_events(headers, records, source_timezone)


def parse_text_playlist_events(
    content: bytes,
    source_timezone: str | None = None,
) -> tuple[dict[str, Any], list[PlaylistEvent]]:
    text = _decode_text(content)
    proteus_result = _parse_proteus_fixed_width(text, source_timezone)
    if proteus_result is not None:
        return proteus_result
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",\t|;")
    except csv.Error as exc:
        raise ValueError(
            "The TXT delimiter could not be detected."
        ) from exc
    reader = csv.DictReader(StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        raise ValueError("The TXT As-Run does not contain headers.")
    headers = [str(header).strip() for header in reader.fieldnames]
    records = [
        {
            str(key).strip(): value
            for key, value in record.items()
            if key is not None
        }
        for record in reader
        if any(str(value or "").strip() for value in record.values())
    ]
    return _structured_events(headers, records, source_timezone)


def _parse_proteus_fixed_width(
    text: str,
    source_timezone: str | None,
) -> tuple[dict[str, Any], list[PlaylistEvent]] | None:
    lines = [
        line.rstrip("\r")
        for line in text.splitlines()
        if line.strip()
    ]
    if not lines:
        return None

    candidates = [
        line
        for line in lines
        if len(line) >= 119
        and line[66:71].strip().lower() in {
            "aired",
            "failed",
            "skipp",
        }
        and _parse_date(line[79:87]) is not None
        and match(r"^\d{2}:\d{2}:\d{2};\d{2}$", line[93:104])
    ]
    if len(candidates) < max(1, int(len(lines) * 0.8)):
        return None

    timezone_info = _timezone(source_timezone)
    events: list[PlaylistEvent] = []
    skipped_statuses = Counter()
    for row_number, line in enumerate(lines, start=1):
        if len(line) < 119:
            continue
        status = line[66:71].strip()
        if status.lower() != "aired":
            skipped_statuses[status or "(blank)"] += 1
            continue
        asset_id = line[0:38].strip()
        air_date = line[79:87].strip()
        timecode = line[93:104].strip()
        duration = line[108:119].strip() or None
        if not asset_id:
            continue
        try:
            air_datetime = datetime.strptime(
                f"{air_date} {timecode[:8]}",
                "%m/%d/%y %H:%M:%S",
            ).replace(tzinfo=timezone_info)
        except ValueError:
            continue
        events.append(PlaylistEvent(
            channel_name=None,
            air_datetime=air_datetime,
            duration=duration,
            asset_id=asset_id,
            source_row=row_number,
        ))

    structure = {
        "format": "Proteus fixed-width As-Run",
        "encoding": "UTF-16",
        "header_row": None,
        "headers": [
            "Asset ID",
            "Status",
            "Air Date",
            "Air Timecode",
            "Duration Timecode",
        ],
        "detected_columns": {
            "asset_id": "Asset ID",
            "time": "Air Timecode",
            "duration": "Duration Timecode",
        },
        "metadata": {
            "date": (
                min(event.air_datetime for event in events).date().isoformat()
                if events
                else None
            ),
            "start_time": (
                min(event.air_datetime for event in events).time().isoformat()
                if events
                else None
            ),
            "channel_name": None,
            "source_timezone": source_timezone,
        },
        "rows": len(lines),
        "asset_occurrences": len(events),
        "unique_assets": len({event.asset_id for event in events}),
        "skipped_statuses": dict(skipped_statuses),
    }
    return structure, events


def parse_xml_playlist_events(
    content: bytes,
    source_timezone: str | None = None,
) -> tuple[dict[str, Any], list[PlaylistEvent]]:
    if not content:
        raise ValueError("The XML As-Run file is empty.")
    if len(content) > MAX_PLAYLIST_SIZE:
        raise ValueError("The XML As-Run file exceeds the 20 MB import limit.")
    try:
        root = ElementTree.fromstring(content)
    except ElementTree.ParseError as exc:
        raise ValueError("The XML As-Run file is invalid.") from exc

    event_nodes = list(root.findall(".//event"))
    if not event_nodes:
        raise ValueError("The XML As-Run does not contain event records.")
    records = [
        {
            child.tag: (child.text or "").strip()
            for child in event
        }
        for event in event_nodes
    ]
    headers = list(dict.fromkeys(
        key
        for record in records
        for key in record
    ))
    return _structured_events(headers, records, source_timezone)


def parse_excel_playlist_events(
    content: bytes,
    source_timezone: str | None = None,
) -> tuple[dict[str, Any], list[PlaylistEvent]]:
    if not content:
        raise ValueError("The As-Run file is empty.")
    if len(content) > MAX_PLAYLIST_SIZE:
        raise ValueError("The As-Run file exceeds the 20 MB import limit.")

    try:
        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ValueError("The Excel As-Run file could not be opened.") from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("The Excel As-Run file does not contain any data.")

    header_index = _detect_header([
        [str(value or "").strip() for value in row]
        for row in rows[:25]
    ])
    headers = [str(value or "").strip() for value in rows[header_index]]
    detected = _detect_columns(headers)
    if not detected["asset_id"] or not detected["time"]:
        raise ValueError(
            "Asset Name and Headend Start Time columns must be mapped."
        )

    timezone_name = None if source_timezone == "auto" else source_timezone
    if source_timezone == "auto":
        raise ValueError(
            "The Excel As-Run does not declare a time zone. "
            "Select the source time zone manually."
        )
    timezone_info = None
    if timezone_name:
        try:
            timezone_info = ZoneInfo(timezone_name)
        except ZoneInfoNotFoundError as exc:
            raise ValueError(
                f'Unknown source time zone "{timezone_name}".'
            ) from exc

    asset_index = headers.index(detected["asset_id"])
    time_index = headers.index(detected["time"])
    duration_index = (
        headers.index(detected["duration"])
        if detected["duration"]
        else None
    )
    events = []
    for row_number, row in enumerate(
        rows[header_index + 1:],
        start=header_index + 2,
    ):
        asset_id = str(row[asset_index] or "").strip()
        air_datetime = _excel_datetime(row[time_index])
        if not asset_id or air_datetime is None:
            continue
        events.append(PlaylistEvent(
            channel_name=None,
            air_datetime=air_datetime.replace(tzinfo=timezone_info),
            duration=(
                _amagi_duration(row[duration_index])
                if duration_index is not None
                else None
            ),
            asset_id=asset_id,
            source_row=row_number,
        ))

    structure = {
        "header_row": header_index + 1,
        "headers": [header for header in headers if header],
        "detected_columns": detected,
        "metadata": {
            "date": (
                min(event.air_datetime for event in events).date().isoformat()
                if events
                else None
            ),
            "start_time": None,
            "channel_name": None,
            "source_timezone": timezone_name,
        },
        "rows": len(rows) - header_index - 1,
        "asset_occurrences": len(events),
        "unique_assets": len({event.asset_id for event in events}),
    }
    return structure, events


def parse_playlist_file(
    filename: str,
    content: bytes,
    source_timezone: str | None = None,
) -> tuple[dict[str, Any], list[PlaylistEvent]]:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        return parse_playlist_events(content, source_timezone)
    if extension == ".xlsx":
        return parse_excel_playlist_events(content, source_timezone)
    if extension == ".json":
        return parse_json_playlist_events(content, source_timezone)
    if extension == ".txt":
        return parse_text_playlist_events(content, source_timezone)
    if extension == ".xml":
        return parse_xml_playlist_events(content, source_timezone)
    raise ValueError(
        "Only .csv, .xlsx, .json, .txt, and .xml As-Run files are supported."
    )


def _filter_values(value: str | None) -> list[str]:
    if not value:
        return []

    return [
        item.strip().lower()
        for item in value.replace("\n", ",").split(",")
        if item.strip()
    ]


def filter_playlist_events(
    events: list[PlaylistEvent],
    filter_mode: str = "all",
    filter_value: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    start_time: str | None = None,
    end_time: str | None = None,
    broadcast_day_start: str = "06:00:00",
) -> list[PlaylistEvent]:
    allowed_modes = {"all", "prefix", "exact", "contains"}
    if filter_mode not in allowed_modes:
        raise ValueError(f"Unsupported filter mode: {filter_mode}.")

    values = _filter_values(filter_value)
    if filter_mode != "all" and not values:
        raise ValueError("At least one filter value is required.")

    parsed_start_date = (
        datetime.strptime(start_date, "%Y-%m-%d").date()
        if start_date
        else None
    )
    parsed_end_date = (
        datetime.strptime(end_date, "%Y-%m-%d").date()
        if end_date
        else None
    )
    parsed_start_time = time.fromisoformat(start_time) if start_time else None
    parsed_end_time = time.fromisoformat(end_time) if end_time else None
    parsed_broadcast_start = time.fromisoformat(broadcast_day_start)

    if (
        parsed_start_date is not None
        and parsed_end_date is not None
        and parsed_end_date < parsed_start_date
    ):
        raise ValueError("End date must be on or after start date.")

    matches: list[PlaylistEvent] = []
    timezone_info = events[0].air_datetime.tzinfo if events else None
    start_boundary = (
        datetime.combine(
            parsed_start_date,
            parsed_broadcast_start,
            tzinfo=timezone_info,
        )
        if parsed_start_date
        else None
    )
    end_boundary = (
        datetime.combine(
            parsed_end_date + timedelta(days=1),
            parsed_broadcast_start,
            tzinfo=timezone_info,
        )
        if parsed_end_date
        else None
    )

    for event in events:
        event_time = event.air_datetime.time()
        asset = event.asset_id.lower()

        if start_boundary and event.air_datetime < start_boundary:
            continue
        if end_boundary and event.air_datetime >= end_boundary:
            continue

        if parsed_start_time and parsed_end_time:
            in_time_range = (
                parsed_start_time <= event_time <= parsed_end_time
                if parsed_start_time <= parsed_end_time
                else event_time >= parsed_start_time
                or event_time <= parsed_end_time
            )
            if not in_time_range:
                continue
        elif parsed_start_time and event_time < parsed_start_time:
            continue
        elif parsed_end_time and event_time > parsed_end_time:
            continue

        if filter_mode == "prefix" and not any(
            asset.startswith(value)
            for value in values
        ):
            continue
        if filter_mode == "exact" and asset not in values:
            continue
        if filter_mode == "contains" and not any(
            value in asset
            for value in values
        ):
            continue

        matches.append(event)

    return sorted(matches, key=lambda event: event.air_datetime)
