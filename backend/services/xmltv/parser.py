import csv
import re
import unicodedata
from datetime import date, datetime, time, timedelta
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from backend.models.programme import Programme


EXPECTED_COLUMNS = [
    "Channel (Optional)",
    "Air Date",
    "Start Time",
    "Program Title",
    "Duration (Conditional)",
    "Parental Rating",
    "Program Description (Conditional)",
    "Original Title (Optional)",
    "Cast (Optional)",
    "Season Number (Optional)",
    "Episode Number (Optional)",
    "Original Episode Title (Optional)",
    "Episode Description (Conditional)",
    "Genre",
    "Country of Production (Optional)",
    "Production Year (Optional)",
    "Premiere (Optional)",
    "Live (Optional)",
    "New (Optional)",
    "Asset ID (Optional)",
    "Original Air Date (Optional)",
    "Icon URL (Optional)",
    "Icon Width (Optional)",
    "Icon Height (Optional)",
    "Keywords (Optional)",
    "Previously Shown (Optional)",
]

LEGACY_COLUMN_ALIASES = {
    column.replace(" (Optional)", "").replace(" (Conditional)", ""): column
    for column in EXPECTED_COLUMNS
    if column.endswith((" (Optional)", " (Conditional)"))
}
LEGACY_COLUMN_ALIASES["Duration (Optional)"] = "Duration (Conditional)"


def clean_text(value: Any) -> str | None:
    if value is None:
        return None

    text = str(value).strip()
    return text or None


def parse_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        converted = from_excel(value)
        if isinstance(converted, datetime):
            return converted.date().isoformat()
        if isinstance(converted, date):
            return converted.isoformat()

    text = clean_text(value)

    if not text:
        raise ValueError("Air Date is required.")

    accepted_formats = (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %I:%M:%S %p",
    )

    for date_format in accepted_formats:
        try:
            return datetime.strptime(text, date_format).date().isoformat()
        except ValueError:
            continue

    raise ValueError("Air Date must use YYYY-MM-DD or MM/DD/YYYY.")


def parse_time(value: Any) -> str:
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0).isoformat()

    if isinstance(value, time):
        return value.replace(microsecond=0).isoformat()

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        converted = from_excel(value)
        if isinstance(converted, datetime):
            return converted.time().replace(microsecond=0).isoformat()
        if isinstance(converted, time):
            return converted.replace(microsecond=0).isoformat()

    text = clean_text(value)

    if not text:
        raise ValueError("Start Time is required.")

    accepted_formats = (
        "%H:%M:%S",
        "%H:%M",
        "%I:%M %p",
        "%I:%M:%S %p",
    )

    for time_format in accepted_formats:
        try:
            return datetime.strptime(text.upper(), time_format).time().isoformat()
        except ValueError:
            continue

    raise ValueError("Start Time is invalid.")


def parse_integer(value: Any) -> int | None:
    text = clean_text(value)

    if text is None:
        return None

    number = int(float(text))

    if number < 0:
        raise ValueError("Numeric values cannot be negative.")

    return number


def parse_boolean(value: Any) -> bool:
    text = clean_text(value)

    if text is None:
        return False

    normalized = text.lower()

    if normalized in {"yes", "true", "1", "y", "sí", "si", "s"}:
        return True

    if normalized in {"no", "false", "0", "n"}:
        return False

    raise ValueError("Use Yes or No.")


def normalize_boolean(
    value: Any,
    field: str,
    source_row: int,
    auto_fixes: list[dict[str, Any]] | None,
) -> bool:
    result = parse_boolean(value)
    text = clean_text(value)

    if text and text not in {"Yes", "No"} and auto_fixes is not None:
        auto_fixes.append({
            "row": source_row,
            "field": field,
            "original_value": text,
            "normalized_value": "Yes" if result else "No",
            "message": f"{field} was normalized to Yes or No.",
        })

    return result


def normalize_duration(
    value: Any,
    source_row: int,
    auto_fixes: list[dict[str, Any]] | None,
) -> str | None:
    if value is None:
        return None

    normalized: str

    if isinstance(value, timedelta):
        total_seconds = int(value.total_seconds())
        normalized = (
            f"{total_seconds // 3600:02d}:"
            f"{(total_seconds % 3600) // 60:02d}:"
            f"{total_seconds % 60:02d}"
        )
    elif isinstance(value, time):
        normalized = value.replace(microsecond=0).isoformat()
    elif (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
    ) or (
        isinstance(value, str)
        and value.strip().replace(".", "", 1).isdigit()
    ):
        numeric_value = float(value)

        if numeric_value <= 0:
            return clean_text(value)

        total_seconds = round(numeric_value * 60)
        normalized = (
            f"{total_seconds // 3600:02d}:"
            f"{(total_seconds % 3600) // 60:02d}:"
            f"{total_seconds % 60:02d}"
        )
    else:
        return clean_text(value)

    if auto_fixes is not None:
        auto_fixes.append({
            "row": source_row,
            "field": "Duration (Conditional)",
            "original_value": str(value),
            "normalized_value": normalized,
            "message": "Numeric duration was interpreted as minutes.",
        })

    return normalized


RATING_ALIASES = {
    "TVY": "TV-Y",
    "TVY7": "TV-Y7",
    "TVG": "TV-G",
    "TVPG": "TV-PG",
    "TV14": "TV-14",
    "TVMA": "TV-MA",
}


def normalize_rating(
    value: Any,
    source_row: int,
    auto_fixes: list[dict[str, Any]] | None,
) -> str | None:
    text = clean_text(value)

    if not text:
        return None

    compact = text.upper().replace("-", "").replace(" ", "")
    normalized = RATING_ALIASES.get(compact, text)

    if normalized != text and auto_fixes is not None:
        auto_fixes.append({
            "row": source_row,
            "field": "Parental Rating",
            "original_value": text,
            "normalized_value": normalized,
            "message": "Parental Rating was normalized.",
        })

    return normalized


def parse_cast(value: Any) -> list[str]:
    text = clean_text(value)

    if not text:
        return []

    return [
        name.strip()
        for name in text.replace(",", ";").split(";")
        if name.strip()
    ]


def parse_keywords(value: Any) -> list[str]:
    text = clean_text(value)

    if not text:
        return []

    return [
        keyword.strip()
        for keyword in text.replace(",", ";").split(";")
        if keyword.strip()
    ]


def parse_optional_date(value: Any) -> str | None:
    if clean_text(value) is None:
        return None
    return parse_date(value)


def generate_asset_id(
    title: str,
    air_date: str,
    season_number: int | None,
    episode_number: int | None,
    episode_title: str | None,
    original_air_date: str | None,
) -> str:
    normalized_title = unicodedata.normalize("NFKD", title)
    slug = re.sub(
        r"[^a-z0-9]+",
        "-",
        normalized_title.encode("ascii", "ignore").decode().lower(),
    ).strip("-") or "programme"

    if season_number is not None and episode_number is not None:
        return f"{slug}-s{season_number:02d}e{episode_number:02d}"

    if episode_title:
        normalized_episode = unicodedata.normalize("NFKD", episode_title)
        episode_slug = re.sub(
            r"[^a-z0-9]+",
            "-",
            normalized_episode.encode("ascii", "ignore").decode().lower(),
        ).strip("-")
        if episode_slug:
            return f"{slug}-{episode_slug}"

    reference_date = original_air_date or air_date
    return f"{slug}-{reference_date.replace('-', '')}"


def read_csv_rows(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HTTPException(
            status_code=400,
            detail="The CSV file must use UTF-8 encoding.",
        ) from exc

    reader = csv.DictReader(StringIO(text))

    if not reader.fieldnames:
        raise HTTPException(
            status_code=400,
            detail="The CSV file does not contain headers.",
        )

    headers = [str(header).strip() for header in reader.fieldnames]
    rows = [
        dict(row)
        for row in reader
        if any(clean_text(value) for value in row.values())
    ]

    return headers, rows


def read_excel_rows(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        workbook = load_workbook(
            filename=BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail="The Excel file could not be opened.",
        ) from exc

    if "Programming" not in workbook.sheetnames:
        raise HTTPException(
            status_code=400,
            detail='The Excel file must contain a sheet named "Programming".',
        )

    worksheet = workbook["Programming"]

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in worksheet[4]
    ]

    rows: list[dict[str, Any]] = []

    for values in worksheet.iter_rows(min_row=5, values_only=True):
        if not any(clean_text(value) for value in values):
            continue

        rows.append(dict(zip(headers, values)))

    return headers, rows


def read_schedule_file(
    filename: str,
    content: bytes,
) -> tuple[list[str], list[dict[str, Any]]]:
    extension = Path(filename).suffix.lower()

    if extension == ".csv":
        headers, rows = read_csv_rows(content)
    elif extension == ".xlsx":
        headers, rows = read_excel_rows(content)
    else:
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx and .csv files are supported.",
        )

    normalized_headers = [
        LEGACY_COLUMN_ALIASES.get(header, header)
        for header in headers
    ]
    normalized_rows = [
        {
            LEGACY_COLUMN_ALIASES.get(key, key): value
            for key, value in row.items()
        }
        for row in rows
    ]
    return normalized_headers, normalized_rows


def build_programme(
    row: dict[str, Any],
    source_row: int,
    auto_fixes: list[dict[str, Any]] | None = None,
) -> Programme:
    title = clean_text(row.get("Program Title"))

    if not title:
        raise ValueError("Program Title is required.")

    air_date = parse_date(row.get("Air Date"))
    season_number = parse_integer(row.get("Season Number (Optional)"))
    episode_number = parse_integer(row.get("Episode Number (Optional)"))
    episode_title = clean_text(row.get("Original Episode Title (Optional)"))
    original_air_date = parse_optional_date(
        row.get("Original Air Date (Optional)")
    )
    asset_id = clean_text(row.get("Asset ID (Optional)")) or generate_asset_id(
        title,
        air_date,
        season_number,
        episode_number,
        episode_title,
        original_air_date,
    )

    return Programme(
        source_row=source_row,
        channel=clean_text(row.get("Channel (Optional)")),
        air_date=air_date,
        start_time=parse_time(row.get("Start Time")),
        program_title=title,
        duration=normalize_duration(
            row.get("Duration (Conditional)"),
            source_row,
            auto_fixes,
        ),
        parental_rating=normalize_rating(
            row.get("Parental Rating"),
            source_row,
            auto_fixes,
        ),
        program_description=clean_text(
            row.get("Program Description (Conditional)")
        ),
        original_title=clean_text(row.get("Original Title (Optional)")),
        cast=parse_cast(row.get("Cast (Optional)")),
        season_number=season_number,
        episode_number=episode_number,
        original_episode_title=episode_title,
        episode_description=clean_text(
            row.get("Episode Description (Conditional)")
        ),
        genre=clean_text(row.get("Genre")),
        country_of_production=clean_text(
            row.get("Country of Production (Optional)")
        ),
        production_year=parse_integer(row.get("Production Year (Optional)")),
        premiere=normalize_boolean(
            row.get("Premiere (Optional)"),
            "Premiere",
            source_row,
            auto_fixes,
        ),
        live=normalize_boolean(
            row.get("Live (Optional)"),
            "Live",
            source_row,
            auto_fixes,
        ),
        new=normalize_boolean(
            row.get("New (Optional)"),
            "New",
            source_row,
            auto_fixes,
        ),
        asset_id=asset_id,
        original_air_date=original_air_date,
        icon_url=clean_text(row.get("Icon URL (Optional)")),
        icon_width=parse_integer(row.get("Icon Width (Optional)")),
        icon_height=parse_integer(row.get("Icon Height (Optional)")),
        keywords=parse_keywords(row.get("Keywords (Optional)")),
        previously_shown=normalize_boolean(
            row.get("Previously Shown (Optional)"),
            "Previously Shown",
            source_row,
            auto_fixes,
        ),
    )
